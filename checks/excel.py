import os
import random

import openpyxl


def read_plates(path: str, sample_size: int | None = None) -> list[str]:
    plate_col = os.environ.get("RENEWAL_PLATE_FIELD", "PLACAS")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_idx = headers.index(plate_col)
    except ValueError:
        raise ValueError(f"Columna '{plate_col}' no encontrada en Excel. Columnas disponibles: {headers}")

    plates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[col_idx]
        if val and str(val).strip():
            plates.append(str(val).strip())

    wb.close()

    if not plates:
        raise ValueError(f"No se encontraron placas en columna '{plate_col}'")

    n = sample_size or int(os.environ.get("QA_SAMPLE_SIZE", "50"))
    n = min(n, 200, len(plates))
    return random.sample(plates, n)
